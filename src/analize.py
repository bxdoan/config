import csv
import openpyxl
from os import path

CODE_HOME = path.abspath(path.dirname(__file__) + '/..')
TMP = f"{CODE_HOME}/tmp"
COLUMN_MAPPING = {
    'HolderAddress': 'address',
    'Balance': 'balance',
    'PendingBalanceUpdate': 'update',
}


class AnaliseConfig(object):

    def __init__(self, **kwargs):
        self.dir = kwargs.get('dir') or ''
        self.list = [40, 35, 30, 25, 20, 15, 10]

    def process(self):
        datas = self._parser_file()

        sum_all, count = self._cal(datas)
        print(f"{sum_all=} and {count=}\n")

        for number in self.list:
            self._cal_and_print(datas=datas, top_address=number, sum_all=sum_all)

    def _cal_and_print(self, datas=None, top_address=0, sum_all=0):
        if not datas:
            datas = []
        sum_top, top_address = self._cal(datas, top_address)
        percent = round(sum_top * 100 / sum_all)
        print(f"{sum_top=} and {top_address=} and {percent=}\n")

    def _cal(self, datas=None, top_address=0):
        if not datas:
            datas = []

        sum_all = 0
        count = 0
        for index, d in enumerate(datas):
            if d.get('balance'):
                value = float(d.get('balance'))
                sum_all += value
                count += 1

            if top_address and index == top_address - 1:
                break

        return round(sum_all), round(count)

    def _parser_file(self):
        parsed_records = []
        raw_v_rows = []

        # region read file upload
        if self.dir.lower().endswith('.csv'):
            raw_v_rows = self._read_csv_file(self.dir, COLUMN_MAPPING)
        elif self.dir.lower().endswith('.xlsx'):
            raw_v_rows = self._read_xlsx_file(self.dir, COLUMN_MAPPING)
        else:
            raise Exception
        # endregion

        # region covert data
        parser_all = {
            'address': lambda v: str(v).strip() if v else None,
            'balance': lambda v: float(str(v).strip()) if v else None,
            'update': lambda v: str(v).strip() if v else None,
        }

        kept_as_is = lambda v: v
        for rvr in raw_v_rows:
            pr = dict()  # pr aka parsed_row
            for k, v in rvr.items():  # :k aka key, :v aka value
                parser_func = parser_all.get(k, kept_as_is)
                pr[k] = parser_func(v)
            parsed_records.append(pr)

        # endregion
        parsed_records = sorted(parsed_records, key=lambda k: k['balance'], reverse=True)
        return parsed_records

    def _read_csv_file(self, file_path, column_mapping):
        raw_v_rows = []  # raw_v_rows aka raw vehicle rows
        # region read csv
        csv.register_dialect('PARCEL_dialect',
                             delimiter=',',
                             quoting=csv.QUOTE_ALL,
                             skipinitialspace=True
                             )
        with open(file_path, mode='r') as csv_file:
            csv_reader = csv.DictReader(csv_file, dialect='PARCEL_dialect')
            for row in csv_reader:
                r = dict()  # r aka record
                for key, value in column_mapping.items():
                    r[value] = row.get(key)
                raw_v_rows.append(r)
        return raw_v_rows

    def _read_xlsx_file(self, file_path, column_mapping):
        wb = openpyxl.load_workbook(file_path)
        first_sheet = wb.sheetnames[0]
        ws = wb[first_sheet]  # the 1st sheet at index 0

        max_column = ws.max_column
        max_row = ws.max_row

        raw_headers = [ws.cell(row=1, column=ci).value for ci in range(1, max_column + 1)]  # ci aka column_index
        raw_headers = list(filter(None, raw_headers))  # remove None column out of header list

        v_fields = [h and column_mapping.get(h.strip()) for h in
                    raw_headers]  # h aka header, ensure header is not null to strip and no error is thrown
        raw_v_rows = []  # raw_v_rows aka raw vehicle rows
        col_count = len(raw_headers)
        for ri in range(2, max_row + 1):  # ri aka row_index - we skip the 1st row which is the header rows
            values = [ws.cell(row=ri, column=ci).value for ci in range(1, col_count + 1)]  # ci aka column_index
            rvr = dict(zip(v_fields, values))  # rvr aka raw_vehicle_row
            raw_v_rows.append(rvr)
        return raw_v_rows


def split_list(list_value=None):
    if list_value is None:
        list_value = []
    if ';' in list_value:
        res = list_value.split(';')
    elif '\n' in list_value:
        res = list_value.split('\n')
    elif '|' in list_value:
        res = list_value.split('|')
    else:
        res = list_value
    return res


if __name__ == '__main__':
    dir_sweat = f"{TMP}/export-sweat.csv"
    dir_people = f"{TMP}/token-people.csv"
    AnaliseConfig(
        dir = dir_people,
    ).process()
